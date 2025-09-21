#!/usr/bin/env python3
"""
Test script to verify Excel hyperlink extraction functionality
"""

import sys
import os
import pandas as pd
import openpyxl

def test_hyperlink_extraction():
    """Test if we can extract hyperlinks from Excel file"""
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'dataset', 'student_details.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found at: {excel_path}")
        return
    
    print(f"Testing Excel file: {excel_path}")
    
    # Test 1: Read with pandas (normal way)
    print("\n=== Testing with pandas ===")
    try:
        df_pandas = pd.read_excel(excel_path)
        print(f"Pandas found {len(df_pandas)} rows")
        print("Columns:", df_pandas.columns.tolist())
        if 'PHOTO' in df_pandas.columns:
            print("Sample PHOTO values:")
            for i, value in enumerate(df_pandas['PHOTO'].head(3)):
                print(f"  Row {i}: {value}")
    except Exception as e:
        print(f"Pandas error: {e}")
    
    # Test 2: Read with openpyxl (hyperlink extraction)
    print("\n=== Testing with openpyxl ===")
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=False)
        worksheet = workbook.active
        print(f"Worksheet name: {worksheet.title}")
        print(f"Max row: {worksheet.max_row}, Max column: {worksheet.max_column}")
        
        # Find header row
        headers = []
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            headers.append(str(cell_value).strip().upper() if cell_value else '')
        
        print("Headers:", headers)
        
        # Find PHOTO column
        photo_col = None
        for col, header in enumerate(headers, 1):
            if 'PHOTO' in header:
                photo_col = col
                break
        
        if photo_col:
            print(f"Found PHOTO column at position {photo_col}")
            print("\nHyperlink extraction test:")
            
            # Extract hyperlinks
            hyperlinks = {}
            for row in range(2, min(worksheet.max_row + 1, 6)):  # Test first 3 data rows
                cell = worksheet.cell(row=row, column=photo_col)
                display_text = str(cell.value) if cell.value else ''
                
                if cell.hyperlink:
                    target = cell.hyperlink.target
                    hyperlinks[row-2] = target  # 0-based index
                    print(f"  Row {row-1}: Display='{display_text}' → Hyperlink='{target}'")
                else:
                    print(f"  Row {row-1}: Display='{display_text}' → No hyperlink")
            
            print(f"\nExtracted {len(hyperlinks)} hyperlinks")
        else:
            print("PHOTO column not found!")
            
    except Exception as e:
        print(f"openpyxl error: {e}")

if __name__ == "__main__":
    test_hyperlink_extraction()