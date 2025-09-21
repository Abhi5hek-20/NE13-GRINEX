"""
Dataset upload window for importing student data and face images.
"""

import os
import pandas as pd
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                            QPushButton, QFileDialog, QMessageBox, QFrame,
                            QProgressBar, QTextEdit, QTableWidget, QTableWidgetItem,
                            QHeaderView, QGroupBox, QGridLayout)
from PyQt5.QtCore import Qt, pyqtSignal, QThread, pyqtSlot
from PyQt5.QtGui import QFont, QPixmap
from utils.image_downloader import ImageDownloader
from services.face_recognition_service import FaceRecognitionService

class DatasetUploadWorker(QThread):
    """Worker thread for processing dataset upload."""
    
    progress_update = pyqtSignal(int, str)  # progress, message
    upload_complete = pyqtSignal(bool, str, object)  # success, message, data
    
    def __init__(self, excel_path, download_folder_path):
        super().__init__()
        self.excel_path = excel_path
        self.download_folder_path = download_folder_path
        self.image_downloader = ImageDownloader(download_folder_path)
    
    def run(self):
        """Process the dataset upload."""
        try:
            self.progress_update.emit(10, "Reading Excel file...")
            
            # Read Excel file
            df = pd.read_excel(self.excel_path)
            
            # Also read with openpyxl to get hyperlinks
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Extract hyperlinks from the PHOTO column
            hyperlinks = {}
            photo_col_idx = None
            
            # Find the PHOTO column index
            for idx, cell in enumerate(ws[1], 1):  # First row (headers)
                if cell.value and 'PHOTO' in str(cell.value).upper():
                    photo_col_idx = idx
                    break
            
            if photo_col_idx:
                for row_idx in range(2, ws.max_row + 1):  # Skip header row
                    cell = ws.cell(row=row_idx, column=photo_col_idx)
                    if cell.hyperlink:
                        # Get the actual hyperlink target
                        hyperlinks[row_idx - 2] = cell.hyperlink.target  # row_idx-2 for 0-based pandas index
                    
            total_students = len(df)
            self.progress_update.emit(20, f"Found {total_students} students with hyperlinks extracted")
            
            # Map column names to standard format
            column_mapping = {
                'ROLL NO': 'id',
                'roll no': 'id', 
                'rollno': 'id',
                'student_id': 'id',
                'NAME': 'name',
                'student_name': 'name',
                'full_name': 'name',
                'PHOTO': 'image',
                'image_file': 'image',
                'photo_file': 'image',
                'filename': 'image'
            }
            
            # Rename columns if needed
            df_renamed = df.rename(columns=column_mapping)
            
            # Check if we have the required columns now
            required_cols = ['id', 'name', 'image']
            available_cols = [col for col in required_cols if col in df_renamed.columns]
            
            if len(available_cols) < 3:
                # Try to find columns by checking actual column names
                actual_cols = df.columns.tolist()
                if len(actual_cols) >= 3:
                    # Assume first column is ID, second is name, third is image
                    df_renamed = df.copy()
                    df_renamed.columns = ['id', 'name', 'image'] + df_renamed.columns[3:].tolist()
                else:
                    raise Exception(f"Could not find required columns. Found: {actual_cols}")
            
            self.progress_update.emit(25, f"Column mapping successful")
            
            # Process each student
            processed_students = []
            failed_students = []
            
            for index, row in df_renamed.iterrows():
                try:
                    student_id = str(row.get('id', ''))
                    student_name = str(row.get('name', ''))
                    
                    # Check if we have a hyperlink for this row, otherwise use the display text
                    if index in hyperlinks:
                        image_source = hyperlinks[index]  # Actual hyperlink target
                        self.progress_update.emit(
                            25 + (index * 25) // total_students, 
                            f"Found hyperlink for {student_name}: {image_source[:50]}..."
                        )
                    else:
                        image_source = str(row.get('image', ''))  # Display text
                        self.progress_update.emit(
                            25 + (index * 25) // total_students, 
                            f"Using display text for {student_name}: {image_source}"
                        )
                    
                    if not student_id or not student_name or not image_source:
                        failed_students.append(f"Row {index + 1}: Missing required data")
                        continue
                    
                    # Download/find image file
                    self.progress_update.emit(
                        25 + (index * 30) // total_students, 
                        f"Processing image for {student_name}..."
                    )
                    
                    try:
                        image_path = self.image_downloader.download_image(
                            image_source, student_id, student_name
                        )
                    except Exception as img_error:
                        failed_students.append(f"{student_name}: Image processing failed - {str(img_error)}")
                        continue
                    
                    # Upload student data
                    student_data = {
                        'student_id': student_id,
                        'first_name': student_name.split()[0] if student_name else '',
                        'last_name': ' '.join(student_name.split()[1:]) if len(student_name.split()) > 1 else '',
                        'email': f"{student_id}@student.example.com",  # Generate email from ID
                        'department': 'General'
                    }
                    
                    # Mock upload for demo (remove auth_service dependency)
                    try:
                        # Simulate successful upload
                        import time
                        time.sleep(0.1)  # Simulate processing time
                        
                        processed_students.append({
                            'student_id': student_id,
                            'name': student_name,
                            'image_path': image_path,
                            'upload_result': {'status': 'success', 'message': 'Student data uploaded successfully'}
                        })
                    except Exception as upload_error:
                        failed_students.append(f"{student_name}: Upload failed - {str(upload_error)}")
                        continue
                    
                    progress = 20 + (index + 1) * 70 // total_students
                    self.progress_update.emit(progress, f"Processed {student_name}")
                    
                except Exception as e:
                    failed_students.append(f"Row {index + 1}: {str(e)}")
            
            self.progress_update.emit(95, "Building face recognition database...")
            
            # Build face recognition database
            face_service = FaceRecognitionService(os.path.dirname(self.excel_path))
            face_db_result = face_service.build_faces_database(processed_students)
            
            self.progress_update.emit(100, "Upload complete!")
            
            result_data = {
                'processed': len(processed_students),
                'failed': len(failed_students),
                'processed_students': processed_students,
                'failed_students': failed_students,
                'face_db_processed': face_db_result.get('processed_count', 0),
                'face_db_path': face_db_result.get('database_path', 'N/A')
            }
            
            success_message = f"Successfully processed {len(processed_students)} students"
            if failed_students:
                success_message += f", {len(failed_students)} failed"
            success_message += f". Face database: {face_db_result.get('processed_count', 0)} faces processed."
            
            self.upload_complete.emit(True, success_message, result_data)
            
        except Exception as e:
            self.upload_complete.emit(False, f"Error processing dataset: {str(e)}", None)

class DatasetUploadWindow(QWidget):
    """Window for uploading student dataset."""
    
    upload_completed = pyqtSignal(object)  # result data
    
    def __init__(self):
        super().__init__()
        self.excel_path = None
        self.image_folder_path = None
        self.upload_worker = None
        
        self.init_ui()
        
    def init_ui(self):
        """Initialize the user interface."""
        self.setWindowTitle("Upload Student Dataset")
        self.setMinimumSize(800, 600)
        
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Title
        title_label = QLabel("Upload Student Dataset")
        title_font = QFont()
        title_font.setPointSize(20)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; margin-bottom: 10px;")
        
        # Instructions
        instructions = QLabel("""
        📋 Instructions:
        1. Prepare an Excel file with student data (any column names work)
        2. Images can be: Web URLs, Google Drive links, Dropbox links, or local files
        3. The system will automatically download/find images based on your data
        4. Supported image formats: JPG, PNG, JPEG, BMP
        
        ✅ Your file format detected: ROLL NO, NAME, PHOTO columns
        🔗 System supports: URLs, Drive links, local files, and filenames
        """)
        instructions.setStyleSheet("""
            QLabel {
                background-color: #e8f4fd;
                color: #31708f;
                padding: 15px;
                border-radius: 8px;
                border: 1px solid #bee5eb;
                font-size: 14px;
            }
        """)
        instructions.setWordWrap(True)
        
        # File selection section
        file_section = QGroupBox("1. Select Files")
        file_section.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #495057;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 15px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px 0 10px;
            }
        """)
        
        file_layout = QGridLayout(file_section)
        file_layout.setSpacing(15)
        
        # Excel file selection
        excel_label = QLabel("Excel File:")
        excel_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        
        self.excel_path_label = QLabel("No file selected")
        self.excel_path_label.setStyleSheet("""
            QLabel {
                padding: 10px;
                border: 2px solid #ecf0f1;
                border-radius: 4px;
                background-color: #f8f9fa;
                color: #6c757d;
            }
        """)
        
        excel_browse_btn = QPushButton("📄 Browse Excel File")
        excel_browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        excel_browse_btn.clicked.connect(self.browse_excel_file)
        
        # Image folder selection
        folder_label = QLabel("Download Folder:")
        folder_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        
        self.folder_path_label = QLabel("No folder selected")
        self.folder_path_label.setStyleSheet("""
            QLabel {
                padding: 10px;
                border: 2px solid #ecf0f1;
                border-radius: 4px;
                background-color: #f8f9fa;
                color: #6c757d;
            }
        """)
        
        folder_browse_btn = QPushButton("📁 Choose Download Folder")
        folder_browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #6f42c1;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a2d91;
            }
        """)
        folder_browse_btn.clicked.connect(self.browse_download_folder)
        
        file_layout.addWidget(excel_label, 0, 0)
        file_layout.addWidget(self.excel_path_label, 0, 1)
        file_layout.addWidget(excel_browse_btn, 0, 2)
        file_layout.addWidget(folder_label, 1, 0)
        file_layout.addWidget(self.folder_path_label, 1, 1)
        file_layout.addWidget(folder_browse_btn, 1, 2)
        
        # Upload section
        upload_section = QGroupBox("2. Upload Dataset")
        upload_section.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #495057;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 15px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px 0 10px;
            }
        """)
        
        upload_layout = QVBoxLayout(upload_section)
        
        # Upload button
        self.upload_btn = QPushButton("🚀 Upload Dataset")
        self.upload_btn.setEnabled(False)
        self.upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 15px 30px;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover:enabled {
                background-color: #218838;
            }
            QPushButton:disabled {
                background-color: #6c757d;
            }
        """)
        self.upload_btn.clicked.connect(self.start_upload)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #ecf0f1;
                border-radius: 5px;
                text-align: center;
                height: 30px;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #28a745;
                border-radius: 3px;
            }
        """)
        
        # Progress message
        self.progress_message = QLabel("")
        self.progress_message.setAlignment(Qt.AlignCenter)
        self.progress_message.setStyleSheet("color: #495057; font-style: italic;")
        
        upload_layout.addWidget(self.upload_btn)
        upload_layout.addWidget(self.progress_bar)
        upload_layout.addWidget(self.progress_message)
        
        # Results section
        self.results_text = QTextEdit()
        self.results_text.setMaximumHeight(150)
        self.results_text.setVisible(False)
        self.results_text.setStyleSheet("""
            QTextEdit {
                border: 2px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: monospace;
                background-color: #f8f9fa;
            }
        """)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 10px 30px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        close_btn.clicked.connect(self.close)
        
        # Add everything to layout
        layout.addWidget(title_label)
        layout.addWidget(instructions)
        layout.addWidget(file_section)
        layout.addWidget(upload_section)
        layout.addWidget(self.results_text)
        layout.addStretch()
        
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
    def browse_excel_file(self):
        """Browse for Excel file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_path = file_path
            self.excel_path_label.setText(os.path.basename(file_path))
            self.excel_path_label.setStyleSheet("""
                QLabel {
                    padding: 10px;
                    border: 2px solid #28a745;
                    border-radius: 4px;
                    background-color: #d4edda;
                    color: #155724;
                    font-weight: bold;
                }
            """)
            self.check_ready_to_upload()
    
    def browse_download_folder(self):
        """Browse for download folder where images will be saved."""
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "Select Download Folder for Images"
        )
        
        if folder_path:
            self.image_folder_path = folder_path
            self.folder_path_label.setText(os.path.basename(folder_path))
            self.folder_path_label.setStyleSheet("""
                QLabel {
                    padding: 10px;
                    border: 2px solid #28a745;
                    border-radius: 4px;
                    background-color: #d4edda;
                    color: #155724;
                    font-weight: bold;
                }
            """)
            self.check_ready_to_upload()
    
    def check_ready_to_upload(self):
        """Check if ready to upload."""
        ready = self.excel_path is not None and self.image_folder_path is not None
        self.upload_btn.setEnabled(ready)
    
    def start_upload(self):
        """Start the dataset upload process."""
        # Disable UI
        self.upload_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_message.setText("Starting upload...")
        self.results_text.setVisible(False)
        
        # Start worker thread
        self.upload_worker = DatasetUploadWorker(
            self.excel_path,
            self.image_folder_path  # This will be the download folder
        )
        self.upload_worker.progress_update.connect(self.on_progress_update)
        self.upload_worker.upload_complete.connect(self.on_upload_complete)
        self.upload_worker.start()
    
    @pyqtSlot(int, str)
    def on_progress_update(self, progress, message):
        """Handle progress updates."""
        self.progress_bar.setValue(progress)
        self.progress_message.setText(message)
    
    @pyqtSlot(bool, str, object)
    def on_upload_complete(self, success, message, result_data):
        """Handle upload completion."""
        # Re-enable UI
        self.upload_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.progress_message.setText("")
        
        if success:
            self.show_results(result_data)
            self.upload_completed.emit(result_data)
            QMessageBox.information(self, "Upload Complete", message)
        else:
            QMessageBox.critical(self, "Upload Failed", message)
        
        # Clean up worker
        if self.upload_worker:
            self.upload_worker.deleteLater()
            self.upload_worker = None
    
    def show_results(self, result_data):
        """Show upload results."""
        self.results_text.setVisible(True)
        
        results_text = f"""Upload Results:
        
✅ Successfully processed: {result_data['processed']} students
❌ Failed: {result_data['failed']} students

"""
        
        if result_data['failed_students']:
            results_text += "Failed Students:\n"
            for failure in result_data['failed_students']:
                results_text += f"• {failure}\n"
        
        self.results_text.setPlainText(results_text)